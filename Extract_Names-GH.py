#Bo Kulbacki 
#April 22, 2024
#Contains functions that automatically identify names from text
import re
import nltk
from nltk.corpus import stopwords
import openpyxl
import csv

# Download the stopwords dataset if you haven't already


# Get a list of English stopwords
stop = set(stopwords.words('english'))

# Now, you can use the stop_words set in your code to filter out common stopwords from text.

#stop = stopwords.words('english')

string = """
example text
"""

def extract_phone_numbers(string):
    r = re.compile(r'(\d{3}[-\.\s]??\d{3}[-\.\s]??\d{4}|\(\d{3}\)\s*\d{3}[-\.\s]??\d{4}|\d{3}[-\.\s]??\d{4})')
    phone_numbers = r.findall(string)
    return [re.sub(r'\D', '', number) for number in phone_numbers]

def extract_email_addresses(string):
    r = re.compile(r'[\w\.-]+@[\w\.-]+')
    return r.findall(string)

def ie_preprocess(document):
    document = ' '.join([i for i in document.split() if i not in stop])
    sentences = nltk.sent_tokenize(document)
    sentences = [nltk.word_tokenize(sent) for sent in sentences]
    sentences = [nltk.pos_tag(sent) for sent in sentences]
    return sentences

def extract_names(document):
    names = []
    sentences = ie_preprocess(document)
    for tagged_sentence in sentences:
        for chunk in nltk.ne_chunk(tagged_sentence):
            if type(chunk) == nltk.tree.Tree:
                if chunk.label() == 'PERSON':
                    names.append(' '.join([c[0] for c in chunk]))
    return names

if __name__ == '__main__':
    # Open the CSV file for reading
    with open('new_dataframe_address_message.csv', mode='r') as csv_file:
        # Create a CSV reader
        csv_reader = csv.DictReader(csv_file)

        # Initialize a list to store the 'message' column values
        messages = []

        # Loop over the rows in the CSV
        for row in csv_reader:
            # Append the 'message' value from each row to the list
            messages.append(row['Message'])



    counter = 0 
    workbook = openpyxl.load_workbook(".xlsx")
    sheet = workbook.active
    for message in messages:
        counter+=1

        #numbers = extract_phone_numbers(message)
        #emails = extract_email_addresses(message)
        names = extract_names(message)


        # Create a new Excel workbook and select the active sheet
        

        # Add a header (optional)
        row_to_append = sheet.max_row + 1

        
        sheet.cell(row=row_to_append, column=1, value=message)
        # Loop through the names and add them to the Excel sheet
        for col_idx, name in enumerate(names, start=1):
            cell = sheet.cell(row=row_to_append, column=col_idx+1)
            cell.value = name

        # Save the workbook to a file
        
        if counter %100 ==0:
            print("COUNT: ", counter)
    workbook.save('final.xlsx')
    print("done")

